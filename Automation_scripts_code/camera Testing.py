

import subprocess
import time
from threading import Thread
from openpyxl import Workbook
workbook = Workbook()
sheet = workbook.active
sheet['A1'] = 'Sr. Number'
sheet['B1'] = 'Test case  Name'
sheet['C1'] = 'Domain'
sheet['D1'] = 'Status'
sheet['E1'] = 'Images Path'
sheet['F1'] = 'Test Case Details'
sheet['F5'] = 'open picture in gallery'
sheet['F6'] = 'Change picture'
# Starting row number
start_row = 2
# Total number of records
num_records = 5
word = 'Camera'
num_print = 5
path = ''
image_path=''
num_prints = 5
for i in range(num_prints):
    sheet.cell(row=i+2, column=3).value = word
for i in range(start_row, start_row + num_records):
    sr_number = i - start_row + 1
    sheet.cell(row=i, column=1).value = sr_number
row=2
col=5
def execute_adb(command):
    subprocess.call(command, shell=True)

def adbUp():
    execute_adb("adb shell am start -n com.android.camera/com.android.camera.CameraActivity > clr ")
    execute_adb("adb logcat -c")
    execute_adb("adb logcat -c")
    t = Thread(target=execute_adb, args=("adb logcat > camera3.txt",))
    t.start()
    #Normal Mode
    time.sleep(5)
    execute_adb("adb shell input keyevent 27")
    time.sleep(5)
    #Beauty mode
    execute_adb("adb shell input tap 280 1310 ")
    time.sleep(1)
    execute_adb("adb shell input keyevent 27")
    #camera timing is 5 seconds
    time.sleep(4)
    execute_adb("adb shell input keyevent 27")
    time.sleep(5)
    execute_adb("adb shell input tap 550 1310 ")
    time.sleep(2)
    execute_adb("adb shell input keyevent 27")
    time.sleep(5)
    execute_adb("adb shell input tap 500 1310 ")
    time.sleep(2)
    execute_adb("adb shell input keyevent 27")
    time.sleep(5)
    execute_adb("adb shell input tap 500 1310 ")
    time.sleep(2)
    execute_adb("adb shell input keyevent 27")
    time.sleep(8)
    execute_adb("adb shell input tap 480 1310 ")
    time.sleep(2)
    execute_adb("adb shell input keyevent 27")
    time.sleep(10)
    print("open Picture in gallery")
    execute_adb("adb shell input tap 100 1400")
    time.sleep(5)
    print("Change Picture")
    execute_adb("adb shell input keyevent 22")
    time.sleep(3)
    execute_adb("adb shell input keyevent 22")
    time.sleep(3)

    execute_adb("adb shell killall -2 logcat > clr")
    print("Fresh output")
    row = 2
    col = 5
    Disrow = 4
    Discol = 6
    testrow = 2
    testcol = 2
    flag=0
    dic={'NormalMode':0,'BeautyMode':0,'FaceUMode':0,'BlurMode':0,'NighAi':0}
    with open('camera3.txt', 'r') as f:
        for line in f:
            if "onShutterButtonClick-----mCameraMode" and 'CameraBeautyShotMode' in line:
                #print("Beauty mode success")
                dic['BeautyMode']+=1
            elif "onShutterButtonClick" and "CameraNormalMode" in line:
                #print("Normal Mode Success")
                dic['NormalMode']+=1
            elif "onShutterButtonClick" and 'CameraFaceUMode' in line:
                #print("AR sticker Mode Success")
                dic['FaceUMode']+=1
            elif "onShutterButtonClick" and 'CameraBlurMode' in line:
                #print("Portrait Mode success")
                dic['BlurMode']+=1
            elif "onShutterButtonClick" and 'CameraNightAiworks' in line:
                #print("Night Mode success")
                dic['NighAi']+=1
            elif 'mLocalFilePath' in line:
                # path = line.split('=')[1]
                # print(line.split('=')[-1])
                sheet.cell(row=row, column=col).value = line.split('=')[-1]
                sheet.cell(row=Disrow, column=Discol).value = line.split('=')[-1] = 'Save picture in gallery'
                row += 1
            elif 'control.captureIntent:2' in line:
                # print(line.split(':') [-1])
                sheet.cell(row=2, column=6).value = 'Camera on click icon'
                # row +=1
            elif 'onPictureTaken end' in line:
                # print(line.split(':') [-1])
                sheet.cell(row=3, column=6).value = 'Take Picture'
                # row +=1
    if flag==5:
        print("All mode are working fine")
    for key, value in dic.items():
        if value>0:
            print(f"{key} is Working fine")
            row = 2
            for key, value in dic.items():
                sheet.cell(row=row, column=2).value = key
                # sheet.cell(row=row, column=2).value = value
                sheet.cell(row=row, column=4).value = 'Pass'
                row += 1
        else:
            print(f"{key} is Working not fine")
            row = 2
            for key, value in dic.items():

                sheet.cell(row=row, column=2).value = key
                # sheet.cell(row=row, column=2).value = value
                sheet.cell(row=row, column=4).value = 'Fail'
                row += 1
        workbook.save('output1.xlsx')
adbUp()
